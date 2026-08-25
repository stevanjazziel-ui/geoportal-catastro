from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import xlrd
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"C:\Users\PC\Downloads\matriz final.xlsx")
DEFAULT_OUTPUT = REPO_ROOT / "visor-matriz-seguimiento-data.js"

MATRIX_SHEET_NAME = "Matriz Inversi\xf3n"
SUMMARY_SHEET_NAME = "Resumen"
CATALOG_SHEET_NAME = "Cat\xe1logos"
GROUP_ROW_INDEX = 2
HEADER_ROW_INDEX = 3
DATA_START_ROW_INDEX = 4

EXCEL_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
SUPPLEMENT_FIELD_MAP = {
    "CERTIFICADO": "MONTO CERTIFICADO ",
    "COMPROMETIDO": "MONTO COMPROMETIDO",
    "DEVENGADO": "MONTO DEVENGADO",
    "PAGADO": "MONTO EJECUTADO",
}
AUTHORITATIVE_SUPPLEMENT_FIELD_MAP = {
    "CODIFICADO": "MONTO CODIFICADO",
    "CERTIFICADO": "MONTO CERTIFICADO ",
    "COMPROMETIDO": "MONTO COMPROMETIDO",
    "DEVENGADO": "MONTO DEVENGADO",
    "PAGADO": "MONTO EJECUTADO",
    "SALDO_POR_CERTIFICAR": "PENDIENTE POR CERTIFICAR ",
    "SALDO_POR_COMPROMETER": "PENDIENTE POR COMPROMETER",
    "SALDO_POR_DEVENGAR": "PENDIENTE POR DEVENGAR",
    "SALDO_POR_PAGAR": "PENDIENTE POR EJECUTAR",
}
FORCE_OVERRIDE_CODES = {
    "73.02.04.2026.2.4.047.202.099.100.001",
    "73.02.35.2026.2.4.047.201.006.200.001",
    "73.04.05.2026.2.4.047.201.099.100.001",
    "73.08.01.2026.2.4.047.201.099.100.001",
}
FORCE_OVERRIDE_FIELD_MAP = {
    "CODIFICADO": "MONTO CODIFICADO",
    "CERTIFICADO": "MONTO CERTIFICADO ",
    "COMPROMETIDO": "MONTO COMPROMETIDO",
    "DEVENGADO": "MONTO DEVENGADO",
    "PAGADO": "MONTO EJECUTADO",
}
DIRECTION_OVERRIDE_FIELD_MAP = {
    "codified": "MONTO CODIFICADO",
    "certified": "MONTO CERTIFICADO ",
    "committed": "MONTO COMPROMETIDO",
    "accrued": "MONTO DEVENGADO",
    "paid": "MONTO EJECUTADO",
    "balanceToCertify": "PENDIENTE POR CERTIFICAR ",
}
FINANCIAL_FIELD_KEYS = [
    "CODIFICADO",
    "CERTIFICADO",
    "COMPROMETIDO",
    "DEVENGADO",
    "PAGADO",
    "SALDO_POR_CERTIFICAR",
    "SALDO_POR_COMPROMETER",
    "SALDO_POR_DEVENGAR",
    "SALDO_POR_PAGAR",
]
SUPPLEMENT_SHEET_FILTERS = {
    "TECNOLOGIA_DE_LA_INFORMACION_XLS": {"Hoja2"},
    "GESTION_ESTRATEGICA_XLS": {"Hoja2"},
    "CULTURA0_XLS": {"Hoja2"},
    "AMBIENTE0_XLS": {"Hoja2"},
}

HIGHLIGHT_FIELD_ALIASES = {
    "direction": "DIRECCION",
    "project": "PROYECTO_PROGRAMA",
    "generalItem": "PARTIDA_GENERAL",
    "specificItem": "PARTIDA_ESPECIFICA",
    "nature": "NATURALEZA",
    "procurementObject": "OBJETO_DE_CONTRATACION",
    "poaDetail": "DETALLE_POA",
    "poaAmount": "MONTO_POA",
    "pacDetail": "DETALLE_PAC",
    "pacAmount": "MONTO_PAC",
    "caseNumber": "N_TRAMITE",
    "procedureType": "TIPO_PROCEDIMIENTO",
    "codified": "CODIFICADO",
    "certified": "CERTIFICADO",
    "committed": "COMPROMETIDO",
    "accrued": "DEVENGADO",
    "paid": "PAGADO",
    "procurementPhase": "FASE_DE_CONTRATACION",
    "processStatus": "ESTADO_DEL_PROCESO",
    "startDate": "FECHA_INICIO_PREPARATORIA",
    "contractAwardDate": "FECHA_ADJUDICACION",
    "contractDate": "FECHA_CONTRATO",
    "contractDays": "PLAZO_CONTRACTUAL_DIAS",
    "contractEndDate": "FECHA_FIN_CONTRACTUAL",
    "daysRemaining": "DIAS_RESTANTES",
    "contractProgress": "AVANCE_EJECUCION_CONTRACTUAL",
    "physicalProgress": "AVANCE_FISICO_SI_APLICA",
    "milestone": "ENTREGABLE_HITO_ACTUAL",
    "owner": "RESPONSABLE_ADMINISTRADOR",
    "contractFinanceGap": "RELACION_ENTRE_EL_AVANCE_DEL_CONTRATO_Y_EL_DEVENGADO",
    "criticalPoint": "PUNTO_CRITICO",
    "alert": "ALERTA",
    "observation": "OBSERVACION_ACCION_CORRECTIVA",
    "commitmentDate": "FECHA_COMPROMISO_DE_ACCION",
    "closingProjectionAccrued": "DEVENGADO_PROYECTADO_AL_CIERRE",
    "closingProjectionPct": "PROYECCION_DE_CIERRE",
    "executionRisk": "RIESGO_DE_NO_EJECUCION",
}


def normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    text = re.sub(r"_+", "_", text)
    return text


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    return clean_string(value) or None


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        return number
    return None


def source_cell_is_fillable(value: Any) -> bool:
    if value in (None, ""):
        return True
    if isinstance(value, (int, float)):
        return False
    text = clean_string(value)
    if text in {"", "$ -"}:
        return True
    return text in EXCEL_ERRORS


def source_cell_is_zero_fillable(value: Any) -> bool:
    if source_cell_is_fillable(value):
        return True
    number = numeric_value(value)
    if number is None:
        return False
    return abs(number) < 1e-9


def row_numeric_field(row_values: list[Any], columns_by_key: dict[str, dict[str, Any]], field_key: str) -> float | None:
    column = columns_by_key.get(field_key)
    if not column:
        return None
    return numeric_value(row_values[column["index"]])


def has_execution_chain_anomaly(row_values: list[Any], columns_by_key: dict[str, dict[str, Any]]) -> bool:
    codified = row_numeric_field(row_values, columns_by_key, "CODIFICADO")
    certified = row_numeric_field(row_values, columns_by_key, "CERTIFICADO")
    committed = row_numeric_field(row_values, columns_by_key, "COMPROMETIDO")
    accrued = row_numeric_field(row_values, columns_by_key, "DEVENGADO")
    paid = row_numeric_field(row_values, columns_by_key, "PAGADO")

    if codified is not None and certified is not None and certified > codified + 1e-9:
        return True
    if certified is not None and committed is not None and committed > certified + 1e-9:
        return True
    if committed is not None and accrued is not None and accrued > committed + 1e-9:
        return True
    if accrued is not None and paid is not None and paid > accrued + 1e-9:
        return True
    return False


def pick_row_indices(worksheet) -> list[int]:
    active_indices: list[int] = []
    for column_index in range(1, worksheet.max_column + 1):
        header_value = clean_string(worksheet.cell(row=HEADER_ROW_INDEX, column=column_index).value)
        if header_value:
            active_indices.append(column_index)
            continue
        has_data = False
        for row_index in range(DATA_START_ROW_INDEX, worksheet.max_row + 1):
            if worksheet.cell(row=row_index, column=column_index).value not in (None, ""):
                has_data = True
                break
        if has_data:
            active_indices.append(column_index)
    return active_indices


def build_columns(worksheet, active_indices: list[int]) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    seen_keys: Counter[str] = Counter()
    last_group = ""
    for column_index in active_indices:
        group_label = clean_string(worksheet.cell(row=GROUP_ROW_INDEX, column=column_index).value)
        if group_label:
            last_group = group_label
        label = clean_string(worksheet.cell(row=HEADER_ROW_INDEX, column=column_index).value)
        base_key = normalize_key(label) or f"COLUMN_{column_index}"
        seen_keys[base_key] += 1
        unique_key = base_key if seen_keys[base_key] == 1 else f"{base_key}_{seen_keys[base_key]}"
        columns.append(
            {
                "index": column_index - 1,
                "columnIndex": column_index,
                "key": unique_key,
                "baseKey": base_key,
                "label": label or f"Columna {column_index}",
                "group": last_group or "Sin grupo",
            }
        )
    return columns


def row_has_content(values: list[Any]) -> bool:
    meaningful = 0
    for value in values:
        if value not in (None, ""):
            meaningful += 1
        if meaningful >= 2:
            return True
    return False


def build_record(row_number: int, row_values: list[Any], columns: list[dict[str, Any]]) -> dict[str, Any]:
    field_map: dict[str, Any] = {}
    error_count = 0
    for column in columns:
        value = serialize_value(row_values[column["index"]])
        field_map[column["key"]] = value
        if isinstance(value, str) and value in EXCEL_ERRORS:
            error_count += 1

    highlights: dict[str, Any] = {}
    for name, field_key in HIGHLIGHT_FIELD_ALIASES.items():
        highlights[name] = field_map.get(field_key)

    search_parts = [
        highlights.get("direction"),
        highlights.get("project"),
        highlights.get("procurementObject"),
        highlights.get("procedureType"),
        highlights.get("procurementPhase"),
        highlights.get("processStatus"),
        highlights.get("alert"),
        highlights.get("executionRisk"),
    ]
    search_text = " ".join(clean_string(part).lower() for part in search_parts if part)

    return {
        "id": f"row-{row_number}",
        "rowNumber": row_number,
        "fields": field_map,
        "highlights": highlights,
        "flags": {
            "hasFormulaError": error_count > 0,
            "formulaErrorCount": error_count,
        },
        "searchText": search_text,
    }


def compact_sheet_grid(worksheet) -> list[list[Any]]:
    last_row = 0
    last_column = 0
    for row in worksheet.iter_rows():
        row_has_value = False
        for cell in row:
            if cell.value not in (None, ""):
                row_has_value = True
                last_column = max(last_column, cell.column)
        if row_has_value:
            last_row = max(last_row, row[0].row)

    rows: list[list[Any]] = []
    for row_index in range(1, last_row + 1):
        row_values = [
            serialize_value(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, last_column + 1)
        ]
        while row_values and row_values[-1] is None:
            row_values.pop()
        rows.append(row_values)
    return rows


def summarize_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    summary_records = [record for record in records if not record.get("syntheticSupplementRow")]
    totals = defaultdict(float)
    directions: dict[str, dict[str, Any]] = {}
    alert_counts: Counter[str] = Counter()
    risk_counts: Counter[str] = Counter()
    phase_counts: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    procedure_counts: Counter[str] = Counter()
    nature_counts: Counter[str] = Counter()
    rows_with_errors = 0

    for record in summary_records:
        highlights = record["highlights"]
        if record["flags"]["hasFormulaError"]:
            rows_with_errors += 1

        codified = numeric_value(highlights.get("codified")) or 0.0
        certified = numeric_value(highlights.get("certified")) or 0.0
        committed = numeric_value(highlights.get("committed")) or 0.0
        accrued = numeric_value(highlights.get("accrued")) or 0.0
        paid = numeric_value(highlights.get("paid")) or 0.0

        totals["codified"] += codified
        totals["certified"] += certified
        totals["committed"] += committed
        totals["accrued"] += accrued
        totals["paid"] += paid

        direction = clean_string(highlights.get("direction")) or "Sin direccion"
        if direction not in directions:
            directions[direction] = {
                "label": direction,
                "count": 0,
                "codified": 0.0,
                "accrued": 0.0,
                "paid": 0.0,
                "highRisk": 0,
                "activeAlerts": 0,
            }
        directions[direction]["count"] += 1
        directions[direction]["codified"] += codified
        directions[direction]["accrued"] += accrued
        directions[direction]["paid"] += paid

        alert = clean_string(highlights.get("alert"))
        risk = clean_string(highlights.get("executionRisk"))
        phase = clean_string(highlights.get("procurementPhase"))
        status = clean_string(highlights.get("processStatus"))
        procedure = clean_string(highlights.get("procedureType"))
        nature = clean_string(highlights.get("nature"))

        if alert:
            alert_counts[alert] += 1
            if alert != "NORMAL":
                directions[direction]["activeAlerts"] += 1
        if risk:
            risk_counts[risk] += 1
            if risk == "ALTO":
                directions[direction]["highRisk"] += 1
        if phase:
            phase_counts[phase] += 1
        if status:
            status_counts[status] += 1
        if procedure:
            procedure_counts[procedure] += 1
        if nature:
            nature_counts[nature] += 1

    total_rows = len(summary_records)
    active_alerts = sum(count for label, count in alert_counts.items() if label and label != "NORMAL")
    high_risk = risk_counts.get("ALTO", 0)

    direction_rows = sorted(
        directions.values(),
        key=lambda item: (item["codified"], item["accrued"], item["count"]),
        reverse=True,
    )
    for direction in direction_rows:
        direction["executionPct"] = direction["accrued"] / direction["codified"] if direction["codified"] else None
        direction["paymentPct"] = direction["paid"] / direction["accrued"] if direction["accrued"] else None

    return {
        "totalRows": total_rows,
        "rowsWithFormulaErrors": rows_with_errors,
        "activeAlerts": active_alerts,
        "highRiskRows": high_risk,
        "totals": {
            "codified": totals["codified"],
            "certified": totals["certified"],
            "committed": totals["committed"],
            "accrued": totals["accrued"],
            "paid": totals["paid"],
        },
        "ratios": {
            "certifiedVsCodified": totals["certified"] / totals["codified"] if totals["codified"] else None,
            "committedVsCertified": totals["committed"] / totals["certified"] if totals["certified"] else None,
            "accruedVsCommitted": totals["accrued"] / totals["committed"] if totals["committed"] else None,
            "paidVsAccrued": totals["paid"] / totals["accrued"] if totals["accrued"] else None,
            "overallExecution": totals["accrued"] / totals["codified"] if totals["codified"] else None,
        },
        "counts": {
            "alerts": dict(alert_counts),
            "risks": dict(risk_counts),
            "phases": dict(phase_counts),
            "statuses": dict(status_counts),
            "procedures": dict(procedure_counts),
            "natures": dict(nature_counts),
        },
        "topDirections": direction_rows[:12],
    }


def discover_supplement_tables(source_path: Path) -> list[tuple[str, int]]:
    workbook = pd.ExcelFile(source_path)
    allowed_sheets = SUPPLEMENT_SHEET_FILTERS.get(normalize_key(source_path.name))
    tables: list[tuple[str, int]] = []
    for sheet_name in workbook.sheet_names:
        if allowed_sheets and sheet_name not in allowed_sheets:
            continue
        preview = pd.read_excel(source_path, sheet_name=sheet_name, header=None, nrows=20)
        for row_index, row in preview.iterrows():
            normalized = [clean_string(value).upper() for value in row.tolist()]
            if "PARTIDA" in normalized and any("MONTO CODIFICADO" in value for value in normalized):
                tables.append((sheet_name, row_index + 1))
                break
    return tables


def load_supplement_dataframe(source_path: Path, sheet_name: str, header_row_index: int) -> pd.DataFrame:
    supplement = pd.read_excel(source_path, sheet_name=sheet_name, header=header_row_index - 1)
    hidden_positions: set[int] = set()
    if source_path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(source_path, formatting_info=True)
        worksheet = workbook.sheet_by_name(sheet_name)
        hidden_positions = {
            excel_row - (header_row_index + 1)
            for excel_row in range(1, worksheet.nrows + 1)
            if excel_row > header_row_index and getattr(worksheet.rowinfo_map.get(excel_row - 1), "hidden", 0)
        }
    elif source_path.suffix.lower() == ".xlsx":
        workbook = load_workbook(source_path, data_only=True)
        worksheet = workbook[sheet_name]
        hidden_positions = {
            excel_row - (header_row_index + 1)
            for excel_row, dimensions in worksheet.row_dimensions.items()
            if excel_row > header_row_index and dimensions.hidden
        }
    if hidden_positions:
        supplement = supplement.iloc[[index for index in range(len(supplement)) if index not in hidden_positions]]
    return supplement.dropna(how="all")


def build_supplement_updates(
    source_paths: list[Path],
    base_codes: list[str],
    zero_fill_paths: set[Path] | None = None,
    force_paths: set[Path] | None = None,
) -> dict[str, dict[str, tuple[float | None, bool, bool, bool]]]:
    counts = Counter(code for code in base_codes if code)
    zero_fill_paths = zero_fill_paths or set()
    force_paths = force_paths or set()

    updates: dict[str, dict[str, tuple[float | None, bool, bool, bool]]] = {}
    for source_path in source_paths:
        resolved_path = source_path.resolve()
        allow_zero_fill = resolved_path in zero_fill_paths
        force_path = resolved_path in force_paths
        for sheet_name, header_row_index in discover_supplement_tables(source_path):
            supplement = load_supplement_dataframe(source_path, sheet_name, header_row_index)
            supplement.columns = [str(column).strip() for column in supplement.columns]
            if "PARTIDA" not in supplement.columns:
                continue
            supplement["PARTIDA"] = supplement["PARTIDA"].fillna("").astype(str).str.strip()
            supplement = supplement[supplement["PARTIDA"].str.count(r"\.") >= 8]
            if force_path:
                supplement = supplement[supplement["PARTIDA"].isin(counts)]
            else:
                unique_codes = {code for code, count in counts.items() if count == 1}
                supplement = supplement[supplement["PARTIDA"].isin(unique_codes)]

            for _, row in supplement.iterrows():
                code = row["PARTIDA"]
                if not code:
                    continue
                payload = updates.setdefault(code, {})
                if force_path:
                    field_map = AUTHORITATIVE_SUPPLEMENT_FIELD_MAP
                elif code in FORCE_OVERRIDE_CODES:
                    field_map = FORCE_OVERRIDE_FIELD_MAP
                else:
                    field_map = SUPPLEMENT_FIELD_MAP
                force_override = force_path or code in FORCE_OVERRIDE_CODES
                for base_field, supplement_field in field_map.items():
                    value = row.get(supplement_field.strip(), row.get(supplement_field))
                    if pd.isna(value):
                        if force_override:
                            payload[base_field] = (None, allow_zero_fill, force_override, True)
                        continue
                    try:
                        number = float(value)
                    except (TypeError, ValueError):
                        if force_override:
                            payload[base_field] = (None, allow_zero_fill, force_override, True)
                        continue
                    if abs(number) < 1e-9 and not allow_zero_fill and not force_override:
                        continue
                    payload[base_field] = (number, allow_zero_fill, force_override, False)
    return updates


def build_authoritative_direction_scope(source_paths: list[Path], base_direction_by_code: dict[str, str]) -> dict[str, set[str]]:
    scope: dict[str, set[str]] = {}
    for source_path in source_paths:
        matched_directions: Counter[str] = Counter()
        matched_codes: list[str] = []
        for sheet_name, header_row_index in discover_supplement_tables(source_path):
            supplement = load_supplement_dataframe(source_path, sheet_name, header_row_index)
            supplement.columns = [str(column).strip() for column in supplement.columns]
            if "PARTIDA" not in supplement.columns:
                continue
            supplement["PARTIDA"] = supplement["PARTIDA"].fillna("").astype(str).str.strip()
            detail_rows = supplement[supplement["PARTIDA"].str.count(r"\.") >= 8].drop_duplicates("PARTIDA")
            for code in detail_rows["PARTIDA"]:
                direction = base_direction_by_code.get(code)
                if direction:
                    matched_directions[direction] += 1
                    matched_codes.append(code)
        if matched_directions:
            direction, _ = matched_directions.most_common(1)[0]
            scope[direction] = {code for code in matched_codes if base_direction_by_code.get(code) == direction}
    return scope


def build_direction_overrides(source_paths: list[Path], base_direction_by_code: dict[str, str]) -> dict[str, dict[str, float]]:
    overrides: dict[str, dict[str, float]] = {}
    for source_path in source_paths:
        matched_directions: Counter[str] = Counter()
        summed_payload: dict[str, float] = defaultdict(float)
        detail_frames: list[pd.DataFrame] = []
        for sheet_name, header_row_index in discover_supplement_tables(source_path):
            supplement = load_supplement_dataframe(source_path, sheet_name, header_row_index)
            supplement.columns = [str(column).strip() for column in supplement.columns]
            if "PARTIDA" not in supplement.columns:
                continue
            supplement["PARTIDA"] = supplement["PARTIDA"].fillna("").astype(str).str.strip()

            detail_rows = supplement[supplement["PARTIDA"].str.count(r"\.") >= 8].copy()
            if not detail_rows.empty:
                detail_frames.append(detail_rows)

        if not detail_frames:
            continue

        detail_rows = pd.concat(detail_frames, ignore_index=True).drop_duplicates("PARTIDA")
        for code in detail_rows["PARTIDA"]:
            direction = base_direction_by_code.get(code)
            if direction:
                matched_directions[direction] += 1

        for _, row in detail_rows.iterrows():
            for target_field, source_field in DIRECTION_OVERRIDE_FIELD_MAP.items():
                value = row.get(source_field.strip(), row.get(source_field))
                if pd.isna(value):
                    continue
                try:
                    summed_payload[target_field] += float(value)
                except (TypeError, ValueError):
                    continue

        if matched_directions and summed_payload:
            direction, _ = matched_directions.most_common(1)[0]
            overrides[direction] = dict(summed_payload)
    return overrides


def build_authoritative_extra_rows(
    source_paths: list[Path],
    base_codes: set[str],
    base_direction_by_code: dict[str, str],
) -> list[dict[str, Any]]:
    extra_rows: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for source_path in source_paths:
        matched_directions: Counter[str] = Counter()
        detail_frames: list[pd.DataFrame] = []
        for sheet_name, header_row_index in discover_supplement_tables(source_path):
            supplement = load_supplement_dataframe(source_path, sheet_name, header_row_index)
            supplement.columns = [str(column).strip() for column in supplement.columns]
            if "PARTIDA" not in supplement.columns:
                continue
            supplement["PARTIDA"] = supplement["PARTIDA"].fillna("").astype(str).str.strip()
            detail_rows = supplement[supplement["PARTIDA"].str.count(r"\.") >= 8].copy()
            if detail_rows.empty:
                continue
            detail_frames.append(detail_rows)
            for code in detail_rows["PARTIDA"]:
                direction = base_direction_by_code.get(code)
                if direction:
                    matched_directions[direction] += 1

        if not detail_frames or not matched_directions:
            continue

        direction, _ = matched_directions.most_common(1)[0]
        detail_rows = pd.concat(detail_frames, ignore_index=True).drop_duplicates("PARTIDA")
        for _, row in detail_rows.iterrows():
            code = clean_string(row.get("PARTIDA"))
            if not code or code in base_codes or code in seen_codes:
                continue

            def get_value(field: str) -> Any:
                return row.get(field.strip(), row.get(field))

            extra_rows.append(
                {
                    "direction": direction,
                    "code": code,
                    "name": clean_string(get_value("NOMBRE")) or "Registro incorporado desde suplemento",
                    "codified": numeric_value(get_value("MONTO CODIFICADO")),
                    "certified": numeric_value(get_value("MONTO CERTIFICADO ")),
                    "committed": numeric_value(get_value("MONTO COMPROMETIDO")),
                    "accrued": numeric_value(get_value("MONTO DEVENGADO")),
                    "paid": numeric_value(get_value("MONTO EJECUTADO")),
                    "balance_to_certify": numeric_value(get_value("PENDIENTE POR CERTIFICAR ")),
                    "balance_to_commit": numeric_value(get_value("PENDIENTE POR COMPROMETER")),
                    "balance_to_accrue": numeric_value(get_value("PENDIENTE POR DEVENGAR")),
                    "balance_to_pay": numeric_value(get_value("PENDIENTE POR EJECUTAR")),
                    "source": source_path.name,
                }
            )
            seen_codes.add(code)
    return extra_rows


def build_payload(
    source_path: Path,
    supplement_paths: list[Path] | None = None,
    zero_fill_paths: set[Path] | None = None,
    force_paths: set[Path] | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(source_path, data_only=True)
    matrix_sheet = workbook[MATRIX_SHEET_NAME]
    summary_sheet = workbook[SUMMARY_SHEET_NAME]
    catalog_sheet = workbook[CATALOG_SHEET_NAME]

    active_indices = pick_row_indices(matrix_sheet)
    columns = build_columns(matrix_sheet, active_indices)
    code_column = next((column for column in columns if column["key"] == "PARTIDA_ESPECIFICA"), None)
    columns_by_key = {column["key"]: column for column in columns}
    base_codes = []
    base_direction_by_code: dict[str, str] = {}
    if code_column:
        direction_column = columns_by_key.get("DIRECCION")
        for row_number in range(DATA_START_ROW_INDEX, matrix_sheet.max_row + 1):
            code = clean_string(matrix_sheet.cell(row=row_number, column=code_column["index"] + 1).value)
            base_codes.append(code)
            if code and direction_column:
                direction = clean_string(matrix_sheet.cell(row=row_number, column=direction_column["index"] + 1).value)
                if direction:
                    base_direction_by_code[code] = direction
    supplement_updates = (
        build_supplement_updates(supplement_paths or [], base_codes, zero_fill_paths, force_paths) if supplement_paths else {}
    )
    direction_overrides = build_direction_overrides(supplement_paths or [], base_direction_by_code) if supplement_paths else {}
    authoritative_direction_scope = (
        build_authoritative_direction_scope(list(force_paths or []), base_direction_by_code) if force_paths else {}
    )
    base_code_counts = Counter(code for code in base_codes if code)
    authoritative_extra_rows = (
        build_authoritative_extra_rows(list(force_paths or []), set(base_code_counts), base_direction_by_code) if force_paths else []
    )
    repeated_code_seen: Counter[str] = Counter()
    records: list[dict[str, Any]] = []

    for row_number in range(DATA_START_ROW_INDEX, matrix_sheet.max_row + 1):
        row_values = [matrix_sheet.cell(row=row_number, column=column_index).value for column_index in range(1, matrix_sheet.max_column + 1)]
        if not row_has_content(row_values):
            continue
        if code_column:
            code = clean_string(row_values[code_column["index"]])
            direction = clean_string(row_values[direction_column["index"]]) if direction_column else ""
            authoritative_codes = authoritative_direction_scope.get(direction)
            if authoritative_codes is not None and code and code not in authoritative_codes:
                for field_key in FINANCIAL_FIELD_KEYS:
                    column = columns_by_key.get(field_key)
                    if column:
                        row_values[column["index"]] = None
            updates = supplement_updates.get(code)
            if updates:
                repeated_code_seen[code] += 1
                is_repeated_authoritative_code = (
                    base_code_counts.get(code, 0) > 1 and any(force_override for _, _, force_override, _ in updates.values())
                )
                execution_anomaly = any(
                    allow_zero_fill for _, allow_zero_fill, _, _ in updates.values()
                ) and has_execution_chain_anomaly(row_values, columns_by_key)
                for field_key, update in updates.items():
                    number, allow_zero_fill, force_override, explicit_clear = update
                    column = columns_by_key.get(field_key)
                    if not column:
                        continue
                    if is_repeated_authoritative_code and repeated_code_seen[code] > 1:
                        row_values[column["index"]] = None
                        continue
                    if force_override:
                        should_fill = True
                    elif allow_zero_fill:
                        should_fill = source_cell_is_zero_fillable(row_values[column["index"]]) or execution_anomaly
                    else:
                        should_fill = source_cell_is_fillable(row_values[column["index"]])
                    if should_fill:
                        row_values[column["index"]] = None if explicit_clear else number
        records.append(build_record(row_number, row_values, columns))

    next_row_number = matrix_sheet.max_row + 1
    for extra_row in authoritative_extra_rows:
        row_values = [None] * matrix_sheet.max_column

        def set_value(key: str, value: Any) -> None:
            column = columns_by_key.get(key)
            if column:
                row_values[column["index"]] = value

        set_value("DIRECCION", extra_row["direction"])
        set_value("PROYECTO_PROGRAMA", extra_row["name"])
        set_value("OBJETO_DE_CONTRATACION", extra_row["name"])
        set_value("PARTIDA_ESPECIFICA", extra_row["code"])
        set_value("CODIFICADO", extra_row["codified"])
        set_value("CERTIFICADO", extra_row["certified"])
        set_value("COMPROMETIDO", extra_row["committed"])
        set_value("DEVENGADO", extra_row["accrued"])
        set_value("PAGADO", extra_row["paid"])
        set_value("SALDO_POR_CERTIFICAR", extra_row["balance_to_certify"])
        set_value("SALDO_POR_COMPROMETER", extra_row["balance_to_commit"])
        set_value("SALDO_POR_DEVENGAR", extra_row["balance_to_accrue"])
        set_value("SALDO_POR_PAGAR", extra_row["balance_to_pay"])
        record = build_record(next_row_number, row_values, columns)
        record["syntheticSupplementRow"] = True
        records.append(record)
        next_row_number += 1

    payload = {
        "meta": {
            "sourceFileName": source_path.name,
            "sourcePath": str(source_path),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sheetNames": workbook.sheetnames,
        },
        "columns": [{key: value for key, value in column.items() if key != "index"} for column in columns],
        "summary": summarize_records(records),
        "records": records,
        "directionOverrides": direction_overrides,
        "sheetViews": [
            {
                "name": SUMMARY_SHEET_NAME,
                "title": "Resumen",
                "description": "Vista de referencia de la hoja de resumen del libro original.",
                "rows": compact_sheet_grid(summary_sheet),
            },
            {
                "name": CATALOG_SHEET_NAME,
                "title": "Catalogos",
                "description": "Contenido auxiliar y explicativo incluido dentro del mismo libro.",
                "rows": compact_sheet_grid(catalog_sheet),
            },
        ],
    }
    return payload


def write_js_payload(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    json_payload = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    output_path.write_text(f"window.MATRIX_VIEWER_DATA = {json_payload};\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a static data bundle for the matriz seguimiento viewer.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Path to the source workbook.")
    parser.add_argument(
        "--supplement",
        type=Path,
        action="append",
        default=None,
        help="Complement workbook to fill financial values by exact partida. Can be passed multiple times.",
    )
    parser.add_argument(
        "--supplement-zero-fill",
        type=Path,
        action="append",
        default=None,
        help="Complement workbook allowed to replace zero values in execution fields by exact partida.",
    )
    parser.add_argument(
        "--supplement-priority",
        type=Path,
        action="append",
        default=None,
        help="Complement workbook that overrides matriz final values by exact partida and direction totals.",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Path to the generated JS data file.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    supplement_paths = list(args.supplement or [])
    zero_fill_paths = list(args.supplement_zero_fill or [])
    priority_paths = list(args.supplement_priority or [])
    ordered_paths: list[Path] = []
    seen_paths: set[Path] = set()
    for path in supplement_paths + zero_fill_paths + priority_paths:
        resolved = path.resolve()
        if resolved in seen_paths:
            continue
        seen_paths.add(resolved)
        ordered_paths.append(path)
    payload = build_payload(
        args.source,
        ordered_paths,
        {path.resolve() for path in zero_fill_paths},
        {path.resolve() for path in priority_paths},
    )
    write_js_payload(payload, args.output)
    print(f"Generated {args.output} from {args.source}")


if __name__ == "__main__":
    main()
