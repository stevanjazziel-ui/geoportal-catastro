from __future__ import annotations

import argparse
import json
import re
import warnings
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


HEADER_ROW = 4
DATA_START_ROW = 5

FIELD_MAP = {
    "Estado": "estado",
    "Nro. trámite": "nroTramite",
    "Código": "codigo",
    "Tipo de trámite": "tipoTramite",
    "Solicitante": "solicitante",
    "Responsable": "responsable",
    "Cargo responsable": "cargoResponsable",
    "Ingresado por": "ingresadoPor",
    "Fecha límite": "fechaLimite",
    "Tiempo restante": "tiempoRestante",
    "Días restantes": "diasRestantes",
    "Prioridad": "prioridad",
    "Fecha de ingreso": "fechaIngreso",
    "Observaciones": "observaciones",
}


def serialize_cell(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value

    return value


def build_payload(source_path: Path) -> dict:
    warnings.filterwarnings(
        "ignore",
        message="Unknown extension is not supported and will be removed",
        module="openpyxl.worksheet._reader",
    )
    warnings.filterwarnings(
        "ignore",
        message="Conditional Formatting extension is not supported and will be removed",
        module="openpyxl.worksheet._reader",
    )

    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    title = sheet["A1"].value or "CONTROL DE TRÁMITES IPRUS"
    source_note = sheet["A2"].value or ""

    header_cells = [sheet.cell(HEADER_ROW, col).value for col in range(1, sheet.max_column + 1)]
    headers = [str(value).strip() if value is not None else "" for value in header_cells]

    unknown_headers = [header for header in headers if header and header not in FIELD_MAP]
    if unknown_headers:
        raise ValueError(f"Se encontraron columnas no mapeadas: {unknown_headers}")

    records = []
    for row_index in range(DATA_START_ROW, sheet.max_row + 1):
        row_values = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "") for value in row_values):
            continue

        item = {}
        for header, value in zip(headers, row_values):
            if not header:
                continue
            item[FIELD_MAP[header]] = serialize_cell(value)

        item["sourceRow"] = row_index
        item["id"] = item.get("codigo") or str(item.get("nroTramite") or row_index)
        records.append(item)

    priorities = {}
    states = {}
    responsibles = {}

    for item in records:
        priorities[item.get("prioridad") or "Sin prioridad"] = priorities.get(item.get("prioridad") or "Sin prioridad", 0) + 1
        states[item.get("estado") or "Sin estado"] = states.get(item.get("estado") or "Sin estado", 0) + 1
        responsibles[item.get("responsable") or "Sin responsable"] = responsibles.get(item.get("responsable") or "Sin responsable", 0) + 1

    source_date = None
    match = re.search(r"(\d{4}-\d{2}-\d{2})", source_note)
    if match:
        source_date = match.group(1)

    return {
        "title": title,
        "sourceFile": source_path.name,
        "sourcePath": str(source_path),
        "sourceNote": source_note,
        "sourceDate": source_date,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "records": records,
        "summary": {
            "total": len(records),
            "priorities": priorities,
            "states": states,
            "responsibles": responsibles,
        },
    }


def write_js_module(payload: dict, output_path: Path) -> None:
    content = "window.TRAMITES_IPRUS_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convierte el Excel de Control de Trámites IPRUS a un archivo JS consumible por el módulo web."
    )
    parser.add_argument(
        "--source",
        default=r"C:\Users\PC\Downloads\Control_Tramites_IPRUS.xlsx",
        help="Ruta del Excel fuente.",
    )
    parser.add_argument(
        "--output",
        default="tramites-iprus-data.js",
        help="Ruta del archivo JS de salida.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_path = Path(args.source).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    payload = build_payload(source_path)
    write_js_module(payload, output_path)
    print(f"Datos generados: {output_path}")
    print(f"Registros procesados: {payload['summary']['total']}")


if __name__ == "__main__":
    main()
