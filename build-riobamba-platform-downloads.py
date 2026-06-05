import argparse
import json
import shutil
import zipfile
from pathlib import Path

import shapefile
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from shapely.geometry import shape as shapely_shape


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "riobamba-censo-data"
MANZANAS_PATH = DATA_DIR / "riobamba_manzanas.geojson"
MANZANAS_STATS_PATH = DATA_DIR / "riobamba_manzanas_stats.json"
PLATFORM_STATS_PATH = DATA_DIR / "riobamba_plataformas_stats.json"
BUS_ISOCHRONES_PATH = DATA_DIR / "riobamba_isocronas_paradas_bus.geojson"
SALUD_ISOCHRONES_PATH = DATA_DIR / "riobamba_isocronas_salud_categorizada.geojson"
OUTPUT_DIR = DATA_DIR / "shp"
MANIFEST_PATH = OUTPUT_DIR / "manifest.json"

PRJ_WGS84 = 'GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]'
BASE_HEADERS = [
    "manzana",
    "plataforma",
    "pob_total",
    "hombres",
    "mujeres",
    "edad0_4",
    "edad5_11",
    "edad12_17",
    "edad18_29",
    "edad30_64",
    "edad65mas",
    "cov_bus",
    "cov_sal",
    "cov_tot",
]
COVERAGE_CONFIGS = [
    {
        "source_path": BUS_ISOCHRONES_PATH,
        "iso_field_prefix": "bus_iso_",
        "label_field_prefix": "bus_nom_",
        "count_field": "cov_bus",
    },
    {
        "source_path": SALUD_ISOCHRONES_PATH,
        "iso_field_prefix": "sal_iso_",
        "label_field_prefix": "sal_nom_",
        "count_field": "cov_sal",
    },
]


def slugify(value: str) -> str:
    normalized = (
        str(value or "")
        .lower()
        .replace("ñ", "enie")
        .replace("Ã±", "enie")
        .replace("Ã‘", "enie")
        .replace("ÃƒÂ±", "enie")
        .replace("Ãƒâ€˜", "enie")
        .replace("Ã¡", "a")
        .replace("Ã©", "e")
        .replace("Ã­", "i")
        .replace("Ã³", "o")
        .replace("Ãº", "u")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    slug = []
    prev_sep = False
    for char in normalized:
        if char.isalnum():
            slug.append(char)
            prev_sep = False
        elif not prev_sep:
            slug.append("_")
            prev_sep = True
    result = "".join(slug).strip("_")
    return result or "seleccion"


def normalize_platform_name(value: str | None) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    return (
        text.replace("Ã‘", "Ñ")
        .replace("Ãƒâ€˜", "Ñ")
        .replace("ÃƒÆ’Ã¢â‚¬Ëœ", "Ñ")
        .replace("Ă‘", "Ñ")
        .replace("Ń", "Ñ")
    )


def is_platform_enie(value: str | None) -> bool:
    return normalize_platform_name(value) == "PLATAFORMA Ñ"


def load_json(path: Path):
    with open(path, "r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def sanitize_token(value):
    token = "".join(char if str(char).isalnum() or char in {"_", "-"} else "_" for char in str(value or "").strip())
    token = "_".join(part for part in token.split("_") if part)
    return token or "sin_nombre"


def build_field_name(prefix: str, index: int) -> str:
    name = f"{prefix}{index}"
    if len(name) <= 10:
        return name
    compact_prefix = prefix.replace("_", "")[:7]
    return f"{compact_prefix}{index}"[:10]


def build_feature_basename(feature, index):
    props = feature.get("properties", {})
    categoria = sanitize_token(props.get("categoria", "sin_categoria")).lower()
    codigo = sanitize_token(props.get("codigo", f"{index:03d}"))
    nombre = sanitize_token(props.get("nombre", f"isocrona_{index:03d}"))[:36]
    return f"iso_{index:03d}_{categoria}_{codigo}_{nombre}"


def build_coverages(source_path: Path):
    payload = load_json(source_path)
    coverages = []
    for index, feature in enumerate(payload.get("features", []), start=1):
        props = feature.get("properties", {})
        coverages.append(
            {
                "iso_name": build_feature_basename(feature, index),
                "label_name": str(props.get("nombre", "") or build_feature_basename(feature, index)),
                "geometry": shapely_shape(feature["geometry"]).buffer(0),
            }
        )
    return coverages


def enrich_plataforma_enie_features(features):
    if not features:
        return features, []

    prepared_coverages = []
    for config in COVERAGE_CONFIGS:
        if not config["source_path"].exists():
            continue
        prepared_coverages.append(
            {
                "iso_field_prefix": config["iso_field_prefix"],
                "label_field_prefix": config["label_field_prefix"],
                "count_field": config["count_field"],
                "coverages": build_coverages(config["source_path"]),
            }
        )

    max_by_group = {item["count_field"]: 0 for item in prepared_coverages}
    enriched = []

    for feature in features:
        rep_point = shapely_shape(feature["geometry"]).representative_point()
        props = dict(feature["properties"])
        total_coverages = 0

        for item in prepared_coverages:
            matching_coverages = sorted(
                (
                    coverage
                    for coverage in item["coverages"]
                    if coverage["geometry"].covers(rep_point)
                ),
                key=lambda coverage: coverage["iso_name"],
            )
            iso_prefix = item["iso_field_prefix"]
            label_prefix = item["label_field_prefix"]
            count_field = item["count_field"]
            props[count_field] = len(matching_coverages)
            total_coverages += len(matching_coverages)
            max_by_group[count_field] = max(max_by_group[count_field], len(matching_coverages))
            for index, coverage in enumerate(matching_coverages, start=1):
                props[build_field_name(iso_prefix, index)] = coverage["iso_name"]
                props[build_field_name(label_prefix, index)] = coverage["label_name"]

        props["cov_tot"] = total_coverages

        enriched.append(
            {
                "type": feature["type"],
                "geometry": feature["geometry"],
                "properties": props,
            }
        )

    extra_headers = []
    for config in COVERAGE_CONFIGS:
        max_overlap = max_by_group.get(config["count_field"], 0)
        extra_headers.extend(build_field_name(config["iso_field_prefix"], index) for index in range(1, max_overlap + 1))
        extra_headers.extend(build_field_name(config["label_field_prefix"], index) for index in range(1, max_overlap + 1))

    return enriched, extra_headers


def feature_parts(feature):
    geometry = feature["geometry"]
    if geometry["type"] == "Polygon":
        return geometry["coordinates"]
    if geometry["type"] == "MultiPolygon":
        parts = []
        for polygon in geometry["coordinates"]:
            parts.extend(polygon)
        return parts
    raise ValueError(f"Geometria no soportada: {geometry['type']}")


def feature_row(feature, extra_headers=None):
    props = feature["properties"]
    row = [
        str(props["man"]),
        str(normalize_platform_name(props["platform_name"]) or "SIN_PLAT"),
        int(props["population_total"]),
        int(props["male"]),
        int(props["female"]),
        int(props["age_0_4"]),
        int(props["age_5_11"]),
        int(props["age_12_17"]),
        int(props["age_18_29"]),
        int(props["age_30_64"]),
        int(props["age_65_plus"]),
        int(props.get("cov_bus", 0) or 0),
        int(props.get("cov_sal", 0) or 0),
        int(props.get("cov_tot", 0) or 0),
    ]
    for header in extra_headers or []:
        row.append(str(props.get(header, "")))
    return row


def write_shapefile_zip(features, basename: str, extra_headers=None):
    temp_dir = OUTPUT_DIR / basename
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)

    shp_base = temp_dir / basename
    writer = shapefile.Writer(str(shp_base), shapeType=shapefile.POLYGON)
    writer.autoBalance = 1

    writer.field("manzana", "C", size=24)
    writer.field("plataforma", "C", size=20)
    writer.field("pob_total", "N", size=12, decimal=0)
    writer.field("hombres", "N", size=12, decimal=0)
    writer.field("mujeres", "N", size=12, decimal=0)
    writer.field("edad0_4", "N", size=12, decimal=0)
    writer.field("edad5_11", "N", size=12, decimal=0)
    writer.field("edad12_17", "N", size=12, decimal=0)
    writer.field("edad18_29", "N", size=12, decimal=0)
    writer.field("edad30_64", "N", size=12, decimal=0)
    writer.field("edad65mas", "N", size=12, decimal=0)
    writer.field("cov_bus", "N", size=12, decimal=0)
    writer.field("cov_sal", "N", size=12, decimal=0)
    writer.field("cov_tot", "N", size=12, decimal=0)
    for header in extra_headers or []:
        writer.field(header[:10], "C", size=120)

    for feature in features:
        writer.poly(feature_parts(feature))
        writer.record(*feature_row(feature, extra_headers=extra_headers))

    writer.close()

    prj_path = shp_base.with_suffix(".prj")
    with open(prj_path, "w", encoding="utf-8") as handle:
        handle.write(PRJ_WGS84)

    zip_path = OUTPUT_DIR / f"{basename}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for extension in (".shp", ".shx", ".dbf", ".prj"):
            file_path = shp_base.with_suffix(extension)
            archive.write(file_path, arcname=file_path.name)

    shutil.rmtree(temp_dir)
    return zip_path


def autosize_columns(ws):
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 24)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_excel(features, basename: str, label: str, extra_headers=None):
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Resumen"
    summary["A1"] = "Seleccion"
    summary["B1"] = label
    summary["A2"] = "Manzanas"
    summary["B2"] = len(features)
    summary["A3"] = "Poblacion total"
    summary["B3"] = sum(int(feature["properties"]["population_total"]) for feature in features)
    summary["A4"] = "Hombres"
    summary["B4"] = sum(int(feature["properties"]["male"]) for feature in features)
    summary["A5"] = "Mujeres"
    summary["B5"] = sum(int(feature["properties"]["female"]) for feature in features)
    for cell in summary["A"]:
        cell.font = Font(bold=True)
    autosize_columns(summary)

    data_sheet = workbook.create_sheet("Datos")
    headers = [*BASE_HEADERS, *(extra_headers or [])]
    data_sheet.append(headers)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True)
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for feature in features:
        data_sheet.append(feature_row(feature, extra_headers=extra_headers))
    autosize_columns(data_sheet)

    xlsx_path = OUTPUT_DIR / f"{basename}.xlsx"
    workbook.save(xlsx_path)
    return xlsx_path


def main():
    parser = argparse.ArgumentParser(description="Genera descargas SHP/XLSX de plataformas de Riobamba.")
    parser.add_argument(
        "--only",
        action="append",
        default=[],
        help="Genera solo la plataforma indicada por basename o etiqueta, por ejemplo plataforma_enie o 'PLATAFORMA Ñ'.",
    )
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    geo = load_json(MANZANAS_PATH)
    stats = load_json(MANZANAS_STATS_PATH)
    platform_stats = load_json(PLATFORM_STATS_PATH)
    existing_manifest = load_json(MANIFEST_PATH) if MANIFEST_PATH.exists() else {}

    by_man = stats.get("byMan", {})
    man_to_platform = platform_stats.get("manToPlatform", {})

    joined_features = []
    for feature in geo["features"]:
        man = feature["properties"]["man"]
        stat = by_man.get(man)
        if not stat:
            continue
        joined_features.append(
            {
                "type": "Feature",
                "geometry": feature["geometry"],
                "properties": {
                    **feature["properties"],
                    **stat,
                    "platform_name": man_to_platform.get(man),
                },
            }
        )

    groups = {
        "riobamba_todas_las_manzanas": {
            "label": "Todas las plataformas",
            "features": joined_features,
        },
        "riobamba_fuera_de_plataformas": {
            "label": "Fuera de plataformas",
            "features": [feature for feature in joined_features if feature["properties"]["platform_name"] is None],
        },
    }

    platform_names = sorted(
        {
            feature["properties"]["platform_name"]
            for feature in joined_features
            if feature["properties"]["platform_name"]
        }
    )
    for platform_name in platform_names:
        groups[slugify(platform_name)] = {
            "label": normalize_platform_name(platform_name),
            "features": [
                feature
                for feature in joined_features
                if feature["properties"]["platform_name"] == platform_name
            ],
        }

    if args.only:
        only_tokens = {slugify(value) for value in args.only}
        groups = {
            basename: group
            for basename, group in groups.items()
            if basename in only_tokens or slugify(group["label"]) in only_tokens
        }

    manifest = {key: value for key, value in existing_manifest.items() if key not in groups}

    for basename, group in groups.items():
        features = group["features"]
        if not features:
            continue

        extra_headers = []
        if basename == "plataforma_enie" or is_platform_enie(group["label"]):
            features, extra_headers = enrich_plataforma_enie_features(features)

        output_label = "PLATAFORMA Ñ" if basename == "plataforma_enie" else group["label"]
        shp_path = write_shapefile_zip(features, basename, extra_headers=extra_headers)
        xlsx_path = write_excel(features, basename, output_label, extra_headers=extra_headers)
        manifest[basename] = {
            "file": shp_path.name,
            "shp_file": shp_path.name,
            "xlsx_file": xlsx_path.name,
            "count": len(features),
            "label": output_label,
        }

    with open(MANIFEST_PATH, "w", encoding="utf-8") as handle:
        json.dump(manifest, handle, ensure_ascii=False, indent=2)

    print("Listo.")
    print(f"Carpeta: {OUTPUT_DIR}")
    print(f"Entradas: {len(manifest)}")


if __name__ == "__main__":
    main()
